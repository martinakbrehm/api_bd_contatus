"""
test_xlsx_exporter.py
----------------------
Testes unitários de xlsx_exporter.gerar_xlsx().

Sem mocks — apenas pandas + openpyxl puro. Verifica que o XLSX gerado
tem a formatação correta (datas, CPF como int, TEL renomeado, etc).
"""

import datetime
import io

import pandas as pd
import pytest
from openpyxl import load_workbook


def _gerar(df: pd.DataFrame):
    from api.utils.xlsx_exporter import gerar_xlsx
    buf = gerar_xlsx(df)
    buf.seek(0)
    wb = load_workbook(buf)
    ws = wb.active
    return ws


def _df_padrao(**kwargs):
    base = {
        "NOME": "JOAO SILVA", "CPF": "09199194996",
        "GENERO": "M", "DATA_NASCIMENTO": "1985-03-15",
        "ENDERECO": "RUA DAS FLORES", "NUM_END": "42",
        "COMPLEMENTO": None, "BAIRRO": "CENTRO",
        "CIDADE": "SAO PAULO", "CEP": "01310100", "UF": "SP",
        "TELEFONE_1": "987654321", "TELEFONE_2": None,
        "TELEFONE_3": None, "TELEFONE_4": None,
        "TELEFONE_5": None, "TELEFONE_6": None,
        "EMAIL_1": "joao@mail.com", "EMAIL_2": None,
    }
    base.update(kwargs)
    return pd.DataFrame([base])


# ── Estrutura do arquivo ──────────────────────────────────────────────────────

class TestEstruturaXlsx:

    def test_retorna_bytes_io(self):
        from api.utils.xlsx_exporter import gerar_xlsx
        buf = gerar_xlsx(_df_padrao())
        assert isinstance(buf, io.BytesIO)
        assert buf.read(4) == b"PK\x03\x04"  # magic bytes XLSX (ZIP)

    def test_tem_aba_lista_pf(self):
        from api.utils.xlsx_exporter import gerar_xlsx
        buf = gerar_xlsx(_df_padrao())
        buf.seek(0)
        wb = load_workbook(buf)
        assert "Lista PF" in wb.sheetnames

    def test_cabecalho_na_primeira_linha(self):
        ws = _gerar(_df_padrao())
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1) if ws.cell(1, c).value]
        assert len(headers) > 0
        assert "NOME" in headers or "CPF" in headers

    def test_dados_na_segunda_linha(self):
        ws = _gerar(_df_padrao())
        # Deve ter pelo menos 2 linhas (cabeçalho + dado)
        assert ws.max_row >= 2

    def test_df_vazio_nao_levanta(self):
        from api.utils.xlsx_exporter import gerar_xlsx
        buf = gerar_xlsx(pd.DataFrame())
        assert isinstance(buf, io.BytesIO)


# ── Renomeação de colunas ─────────────────────────────────────────────────────

class TestRenomeacaoColunas:

    def test_telefone_renomeado_para_tel(self):
        ws = _gerar(_df_padrao())
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        assert "TEL_1" in headers
        assert "TELEFONE_1" not in headers

    def test_colunas_internas_removidas(self):
        df = _df_padrao()
        df["_ID_MAILING"] = 999
        df["_ID_COMPLEMENT"] = 1
        ws = _gerar(df)
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        assert "_ID_MAILING" not in headers
        assert "_ID_COMPLEMENT" not in headers


# ── Formatação de DATA_NASCIMENTO ─────────────────────────────────────────────

class TestDataNascimento:

    def test_data_string_iso_convertida_para_br(self):
        ws = _gerar(_df_padrao(DATA_NASCIMENTO="1985-03-15"))
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        col_idx = headers.index("DATA_NASCIMENTO") + 1
        val = ws.cell(2, col_idx).value
        assert val == "15/03/1985"

    def test_data_objeto_date_convertida(self):
        ws = _gerar(_df_padrao(DATA_NASCIMENTO=datetime.date(1990, 7, 4)))
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        col_idx = headers.index("DATA_NASCIMENTO") + 1
        val = ws.cell(2, col_idx).value
        assert val == "04/07/1990"

    def test_data_none_fica_none(self):
        ws = _gerar(_df_padrao(DATA_NASCIMENTO=None))
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        col_idx = headers.index("DATA_NASCIMENTO") + 1
        val = ws.cell(2, col_idx).value
        assert val is None or str(val).strip() in ("", "None", "nan")


# ── CPF como inteiro ──────────────────────────────────────────────────────────

class TestCpfInt:

    def test_cpf_armazenado_como_inteiro(self):
        """CPF deve ser int para que number_format funcione sem flag verde."""
        ws = _gerar(_df_padrao(CPF="09199194996"))
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        col_idx = headers.index("CPF") + 1
        val = ws.cell(2, col_idx).value
        assert isinstance(val, int)
        assert val == 9199194996

    def test_cpf_nulo_armazenado_como_none(self):
        ws = _gerar(_df_padrao(CPF=None))
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        col_idx = headers.index("CPF") + 1
        assert ws.cell(2, col_idx).value is None


# ── Formatação numérica ───────────────────────────────────────────────────────

class TestFormatoNumerico:

    def test_cel_cpf_tem_number_format(self):
        ws = _gerar(_df_padrao())
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        col_idx = headers.index("CPF") + 1
        assert ws.cell(2, col_idx).number_format == "00000000000"

    def test_cel_tel_tem_number_format(self):
        ws = _gerar(_df_padrao())
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        tel_cols = [i + 1 for i, h in enumerate(headers) if h and h.startswith("TEL_")]
        if tel_cols:
            assert ws.cell(2, tel_cols[0]).number_format == "000000000"


# ── Sanitização ───────────────────────────────────────────────────────────────

class TestSanitizacao:

    def test_caracteres_controle_removidos(self):
        ws = _gerar(_df_padrao(NOME="JOAO\x00SILVA"))
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        col_idx = headers.index("NOME") + 1
        val = ws.cell(2, col_idx).value
        assert "\x00" not in str(val)

    def test_multiplos_registros_preservados(self):
        df = pd.DataFrame([
            _df_padrao(CPF="09199194996").iloc[0].to_dict(),
            _df_padrao(CPF="00000000191").iloc[0].to_dict(),
        ])
        ws = _gerar(df)
        assert ws.max_row == 3  # 1 cabeçalho + 2 linhas
