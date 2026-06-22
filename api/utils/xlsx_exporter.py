"""
xlsx_exporter.py
----------------
Geração de planilha .xlsx formatada a partir de um DataFrame.

Formatação baseada na macro VBA original (xlsx/generator.py):
  - Cabeçalho: negrito, branco sobre roxo escuro (#4A148C), centralizado
  - CPF, NOME, GENERO, UF, DATA_NASCIMENTO: centralizado
  - TELEFONE_*: centralizado, number_format 9 dígitos
  - DDD_*: centralizado, number_format 2 dígitos
  - ENDERECO, BAIRRO, CIDADE, EMAIL_*, CBO: alinhado à esquerda
  - Largura de coluna automática (máx 60)
  - Zoom 90%, cabeçalho congelado (freeze_panes A2)

Produz uma aba:
  Lista PF — registros com a formatação acima
"""

from __future__ import annotations

import datetime
import io

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# (cores de cabeçalho removidas — usa estilo Excel nativo "Input" / "Entrada")

# ── Regras de formatação por coluna ─────────────────────────────────
#   Cada entrada: (alinhamento_horizontal, number_format | None)
_REGRAS_COLUNAS: dict[str, tuple[str, str | None]] = {
    "NOME":            ("left",   None),
    "CPF":             ("center", None),
    "GENERO":          ("center", None),
    "UF":              ("center", None),
    "DATA_NASCIMENTO": ("center", None),
    "CEP":             ("center", "00000000"),
    "ENDERECO":        ("left",   None),
    "BAIRRO":          ("left",   None),
    "CIDADE":          ("left",   None),
    "ATIVIDADE":       ("left",   None),   # profissão (JOIN CBO)
}
_PREFIXO_REGRAS: dict[str, tuple[str, str | None]] = {
    "DDD_":      ("center", "00"),
    "TELEFONE_": ("center", "000000000"),
    "TEL_":      ("center", "000000000"),
    "EMAIL_":    ("left",   None),
}
# Colunas cujo formato é puramente numérico (zeros) — valor deve ser int
_FORMATOS_NUMERICOS: set[str] = {
    fmt for _, fmt in list(_REGRAS_COLUNAS.values()) + list(_PREFIXO_REGRAS.values())
    if fmt and all(c == "0" for c in fmt)
}


def _para_int(val) -> int | None:
    """Converte para int, retorna None se vazio/inválido."""
    if val is None:
        return None
    s = str(val).strip()
    if s in ("", "None", "nan", "NaT", "0"):
        return None
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return None


_PADRAO = ("center", None)


def _regra_coluna(nome: str) -> tuple[str, str | None]:
    if nome in _REGRAS_COLUNAS:
        return _REGRAS_COLUNAS[nome]
    for prefixo, regra in _PREFIXO_REGRAS.items():
        if nome.startswith(prefixo):
            return regra
    return _PADRAO


# Mapeamento de nomes de coluna para exibição no XLSX
_RENAME_DISPLAY = {f"TELEFONE_{i}": f"TEL_{i}" for i in range(1, 7)}


def gerar_xlsx(df: pd.DataFrame, resumo: dict | None = None) -> io.BytesIO:
    """
    Gera um arquivo .xlsx em memória com os registros formatados.

    Parâmetros
    ----------
    df     : DataFrame com os registros a exportar.
    resumo : ignorado (mantido para compatibilidade de chamadas existentes).

    Retorna
    -------
    BytesIO com o conteúdo do arquivo .xlsx pronto para envio via send_file.
    """
    wb = Workbook()

    # ── Preparar DataFrame para exibição ─────────────────────────────────
    # Remove colunas internas de cursor
    df = df.drop(columns=[c for c in df.columns if c.startswith("_")], errors="ignore")

    # Reordena colunas na ordem correta de saída
    from api.utils.data_processor import colunas_saida as _colunas_saida
    com_atividade = "ATIVIDADE" in df.columns
    _ordem = _colunas_saida(com_atividade=com_atividade)
    _cols = [c for c in _ordem if c in df.columns]
    _extras = [c for c in df.columns if c not in _cols]
    df = df[_cols + _extras]

    # Embaralha linhas
    df = df.sample(frac=1).reset_index(drop=True)
    # Renomear TELEFONE_N → TEL_N
    df.rename(columns=_RENAME_DISPLAY, inplace=True)

    # ── Aba "Lista PF" ───────────────────────────────────────────────
    ws = wb.active
    ws.title = "Lista PF"

    headers = df.columns.tolist()
    header_font  = Font(name="Aptos Narrow", size=11)
    data_font    = Font(name="Aptos Narrow", size=11)
    header_align = Alignment(horizontal="center", vertical="center")

    for ci, col in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.style     = "Input"        # estilo "Entrada" do Excel
        cell.font      = header_font    # sobrescreve fonte
        cell.alignment = header_align   # sobrescreve alinhamento

    for ri, row in enumerate(df.itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            col_name = headers[ci - 1]
            alinhamento, fmt = _regra_coluna(col_name)

            # Sanitiza strings: remove caracteres de controle XML-inválidos
            if isinstance(val, str):
                val = "".join(c for c in val if c >= " " or c in "\t\n\r")

            # DATA_NASCIMENTO: converte para string no formato DD/MM/YYYY
            if col_name == "DATA_NASCIMENTO" and val is not None and str(val).strip() not in ("", "None", "nan", "NaT", "NaN"):
                if isinstance(val, (datetime.date, datetime.datetime)):
                    val = val.strftime("%d/%m/%Y")
                elif isinstance(val, str) and len(val) >= 10:
                    try:
                        val = datetime.datetime.strptime(val[:10], "%Y-%m-%d").strftime("%d/%m/%Y")
                    except ValueError:
                        pass

            # Converte para int colunas com formato numérico (evita flag verde)
            if fmt in _FORMATOS_NUMERICOS:
                val = _para_int(val)

            # CPF e NUM_END: extrai dígitos e converte para int (sem flag verde)
            if col_name in ("CPF", "NUM_END"):
                digits = "".join(c for c in str(val) if c.isdigit()) if val is not None else ""
                val = int(digits) if digits else None

            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = data_font
            cell.alignment = Alignment(horizontal=alinhamento, vertical="bottom")
            if col_name == "CPF":
                cell.number_format = "00000000000"
            elif fmt:
                cell.number_format = fmt

    # Largura automática por coluna
    for ci, col in enumerate(headers, 1):
        letra = get_column_letter(ci)
        max_len = max(
            [len(str(col))] + [len(str(v)) for v in df.iloc[:, ci - 1]]
        )
        ws.column_dimensions[letra].width = min(max_len + 2, 60)

    ws.sheet_view.zoomScale = 90
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def gerar_excel_bytes(df: pd.DataFrame) -> io.BytesIO:
    """Alias de gerar_xlsx. Mantido para compatibilidade."""
    return gerar_xlsx(df)
